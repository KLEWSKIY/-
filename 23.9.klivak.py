import os
import re
from openpyxl import load_workbook
def search_excel_files(directory, regex_pattern):
    result = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(root, file)

                workbook = load_workbook(file_path)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    for row in sheet.iter_rows():
                        for cell in row:
                            match = re.search(regex_pattern, str(cell.value))
                            if match:
                                result.append({
                                    'file_path': file_path,
                                    'sheet_name': sheet_name,
                                    'cell_address': cell.coordinate,
                                    'matched_text': match.group()
                                })


                workbook.close()

    return result

if __name__ == "__main__":
    search_directory = r"C:\need\kchay"
    regex_pattern = r'Hello World!'

    search_result = search_excel_files(search_directory, regex_pattern)

    for item in search_result:
        print(f"Файл: {item['file_path']}")
        print(f"Аркуш: {item['sheet_name']}")
        print(f"Адреса клітинки: {item['cell_address']}")
        print(f"Відповідний текст: {item['matched_text']}")
        print("------")
